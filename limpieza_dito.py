import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# --------------------------------------------
# 1. Cargar archivo Excel y hojas necesarias
# --------------------------------------------
archivo = 'assessment powerbi excel.xlsx'
df_ventas = pd.read_excel(archivo, sheet_name='Ventas')
df_costos = pd.read_excel(archivo, sheet_name='Costos')
df_clientes = pd.read_excel(archivo, sheet_name='Clientes')

# --------------------------------------------------
# 🧹 2. Corregir errores ortográficos en columna Zona
# --------------------------------------------------
df_ventas['Zona'] = df_ventas['Zona'].replace({
    'Surrr': 'Sur',
    'Nortee': 'Norte',
    'Centroo': 'Centro'
})

# ---------------------------------------------------------------------
# 3. Limpieza profunda y conversión robusta de columna 'Fecha'
# ---------------------------------------------------------------------
def convertir_fecha(valor):
    try:
        # Si es número tipo Excel serial
        if isinstance(valor, (int, float)):
            return pd.to_datetime('1899-12-30') + pd.to_timedelta(valor, unit='D')
        # Si es texto con símbolos extraños
        elif isinstance(valor, str):
            valor = re.sub(r"[^\d/]", "", valor.strip())
            return pd.to_datetime(valor, dayfirst=True, errors='coerce')
    except:
        return pd.NaT

df_ventas['Fecha'] = df_ventas['Fecha'].apply(convertir_fecha)

# ----------------------------------------------------------------------
# 4. Revisar si quedaron fechas no válidas (NaT después de la conversión)
# ----------------------------------------------------------------------
print("Filas con fechas inválidas:")
print(df_ventas[df_ventas['Fecha'].isna()])

# ------------------------------------------------------------------
# 5. Verificar si existen duplicados exactos y eliminarlos si hay
# ------------------------------------------------------------------
duplicados_exactos = df_ventas[df_ventas.duplicated()]
if not duplicados_exactos.empty:
    print("Duplicados exactos encontrados:\n", duplicados_exactos)
    print("Total duplicados exactos:", len(duplicados_exactos))
    df_ventas = df_ventas.drop_duplicates()

# ---------------------------------------------------
# 6. Formatear la fecha al estilo DD/MM/YYYY (texto final para Excel o Power BI)
# ---------------------------------------------------
df_ventas['Fecha'] = df_ventas['Fecha'].dt.strftime('%d/%m/%Y')

pd.set_option('display.max_rows', None)     # Mostrar todas las filas
pd.set_option('display.max_columns', None)  # Mostrar todas las columnas

#7. Total ventas
df_ventas['Total ventas']=df_ventas['Unidades'] * df_ventas['Precio Unitario']


#8. añade la columna Costo Unitario a df_ventas, buscando coincidencia por Producto.
df_ventas = pd.merge(df_ventas,df_costos[['Producto','Costo Unitario']],on='Producto',how='left')

#9. Calcular Costo total y Margen bruto
df_ventas['Costo total'] = df_ventas['Unidades'] * df_ventas['Costo Unitario']
df_ventas['Margen bruto'] = df_ventas['Total ventas'] - df_ventas['Costo total']

#10.Convertir a tabla estructurada

def convertir_a_tabla_excel(ruta_archivo,nombre_tabla):
    wb=load_workbook(ruta_archivo)
    ws=wb.active

    max_row=ws.max_row
    max_col=ws.max_column

    col_final=get_column_letter(max_col)

    rango=f"A1:{col_final}{max_row}"

    Tabla=Table(displayName=nombre_tabla,ref=rango)
    estilo=TableStyleInfo(name="TableStyleMedium9",showColumnStripes=True)
    Tabla.tableStyleInfo=estilo
    ws.add_table(Tabla)

    wb.save(ruta_archivo)
    print(f"Tabla '{nombre_tabla}'creada en {ruta_archivo}")

# ---------------------------------------------------
# Exportar a Excel para validar visualmente
# ---------------------------------------------------
df_ventas.to_excel("ventas_limpias.xlsx", index=False)
convertir_a_tabla_excel("ventas_limpias.xlsx","Archivo_ventas")

#print(df_ventas)

# ---------------------------------------------------
# 💡 Extras para depuración (descomentar si se desea)
# ---------------------------------------------------
#print(df_ventas.head())
#print(df_costos.head())
#print(df_clientes.head())
#print(df_ventas.columns)
